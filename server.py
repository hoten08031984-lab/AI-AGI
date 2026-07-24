import http.server
import socketserver
import os
import time
import json
import sys
import threading
import unicodedata
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def clean_str(val):
    if val is None:
        return ''
    s = str(val).strip()
    return unicodedata.normalize('NFC', s)

def normalize_loai_cp(val):
    s = clean_str(val)
    if not s or s.lower() == 'none':
        return 'Chưa phân loại'
    s_lower = s.lower()
    # Unify all typos & accent variants of Hành chính phí
    if 'hành' in s_lower or 'hanh' in s_lower or 'hà' in s_lower and 'chí' in s_lower or 'hà' in s_lower and 'chi' in s_lower:
        if 'chí' in s_lower or 'chi' in s_lower:
            return 'Hành chính phí'
    if 'công tác' in s_lower or 'cong tac' in s_lower or 'cô' in s_lower and 'tá' in s_lower:
        return 'Công tác phí'
    if 'bảo trì' in s_lower or 'bao tri' in s_lower:
        return 'Phí bảo trì'
    return s


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PORT = 8080
DIRECTORY = BASE_DIR
EXCEL_PATH = os.path.join(BASE_DIR, "THEO DOI HOP DONG-2_Optimized.xlsx")
JS_DATA_PATH = os.path.join(BASE_DIR, "dashboard_data.js")
JSON_DATA_PATH = os.path.join(BASE_DIR, "dashboard_data.json")

last_mtime = 0
last_check_time = 0
extract_lock = threading.Lock()

def extract_excel_data(force=False):
    global last_mtime, last_check_time
    
    now = time.time()
    if not force and now - last_check_time < 2:
        return False

    if not extract_lock.acquire(blocking=False):
        return False
        
    try:
        last_check_time = now
        if not os.path.exists(EXCEL_PATH):
            print(f"File Excel not found: {EXCEL_PATH}")
            return False
        
        current_mtime = os.path.getmtime(EXCEL_PATH)
        if not force and current_mtime <= last_mtime:
            return False  # No change

        print(f"[{time.strftime('%H:%M:%S')}] Phát hiện kiểm tra Excel (Force={force})! Đang đọc và cập nhật dữ liệu...")
        
        vals = None
        abs_src = os.path.abspath(EXCEL_PATH)

        try:
            import pythoncom
            import win32com.client as win32

            for attempt in range(5):
                pythoncom.CoInitialize()
                excel = None
                wb = None
                try:
                    try:
                        excel = win32.GetActiveObject('Excel.Application')
                    except:
                        excel = win32.DispatchEx('Excel.Application')
                        excel.Visible = False
                        excel.DisplayAlerts = False
                        excel.AutomationSecurity = 1

                    wb_found = None
                    try:
                        for open_wb in excel.Workbooks:
                            if open_wb.FullName.lower() == abs_src.lower():
                                wb_found = open_wb
                                break
                    except:
                        pass

                    if wb_found:
                        wb = wb_found
                        close_wb = False
                    else:
                        wb = excel.Workbooks.Open(abs_src, 0, True)
                        close_wb = True

                    time.sleep(0.5)
                    ws = wb.Sheets('CHI')
                    vals = ws.UsedRange.Value
                    if close_wb:
                        try: wb.Close(False)
                        except: pass
                    break
                except Exception as e_attempt:
                    print(f"Lần thử {attempt+1} đọc Excel bị hoãn: {e_attempt}")
                    time.sleep(1.0)
                finally:
                    pythoncom.CoUninitialize()
        except Exception as e_com:
            print(f"Lỗi win32com: {e_com}, đang dùng openpyxl...")
            try:
                wb_ox = openpyxl.load_workbook(abs_src, data_only=True)
                if 'CHI' in wb_ox.sheetnames:
                    ws_ox = wb_ox['CHI']
                    vals = [list(row) for row in ws_ox.iter_rows(values_only=True)]
            except Exception as e_ox:
                print(f"Lỗi openpyxl: {e_ox}")
                return False


        if not vals:
            return False

        rows = [list(r) for r in vals if r and any(x is not None for x in r)]
        data = rows[1:]
            
        processed_items = []
        for idx, r in enumerate(data):
            if not r or not any(x is not None for x in r):
                continue

            r_padded = list(r) + [None] * (17 - len(r))

            raw_loai_cp = r_padded[0]
            loai_cp = normalize_loai_cp(raw_loai_cp)

            tieu_muc = clean_str(r_padded[1])
            so_hd = clean_str(r_padded[2])
            ngay_hd = clean_str(r_padded[3])[:10]
            thang = clean_str(r_padded[4])
            ly_do = clean_str(r_padded[5])
            chi_tiet = clean_str(r_padded[6])
            chi_tiet_hd = clean_str(r_padded[7])
            kho = clean_str(r_padded[8])

            try: st_no_vat = float(r_padded[9]) if r_padded[9] is not None else 0.0
            except: st_no_vat = 0.0

            try: vat = float(r_padded[10]) if r_padded[10] is not None else 0.0
            except: vat = 0.0

            try: st_vat = float(r_padded[11]) if r_padded[11] is not None else 0.0
            except: st_vat = 0.0

            ngay_tt = clean_str(r_padded[12])[:10]
            nguoi_thu_huong = clean_str(r_padded[13])
            stk = clean_str(r_padded[14])
            ngan_hang = clean_str(r_padded[15])

            try:
                parsed_nam = int(float(r_padded[16])) if r_padded[16] is not None else 0
                if parsed_nam <= 1900 or parsed_nam > 2100:
                    nam = 'N/A'
                else:
                    nam = parsed_nam
            except:
                nam = 'N/A'

            # Ignore only completely empty noise rows
            if loai_cp == 'Chưa phân loại' and not tieu_muc and not so_hd and not ly_do and not chi_tiet and st_vat == 0 and st_no_vat == 0:
                continue

            processed_items.append({
                'id': len(processed_items) + 1,
                'loai_cp': loai_cp,
                'tieu_muc': tieu_muc,
                'so_hd': so_hd,
                'ngay_hd': ngay_hd,
                'thang': thang,
                'ly_do': ly_do,
                'chi_tiet': chi_tiet,
                'chi_tiet_hd': chi_tiet_hd,
                'kho': kho if kho and kho != 'None' else 'Khác',
                'st_no_vat': st_no_vat,
                'vat': vat,
                'st_vat': st_vat,
                'ngay_tt': ngay_tt,
                'nguoi_thu_huong': nguoi_thu_huong if nguoi_thu_huong and nguoi_thu_huong != 'None' else '',
                'stk': stk if stk and stk != 'None' else '',
                'ngan_hang': ngan_hang if ngan_hang and ngan_hang != 'None' else '',
                'nam': nam
            })
            
        sync_info = {
            'last_updated': time.strftime('%Y-%m-%d %H:%M:%S'),
            'total_rows': len(processed_items)
        }
        
        with open(JSON_DATA_PATH, 'w', encoding='utf-8') as f:
            json.dump(processed_items, f, ensure_ascii=False, indent=2)

        with open(JS_DATA_PATH, 'w', encoding='utf-8') as f:
            f.write(f'window.SYNC_INFO = {json.dumps(sync_info, ensure_ascii=False)};\n')
            f.write('window.RAW_DATA = ' + json.dumps(processed_items, ensure_ascii=False) + ';')

        last_mtime = current_mtime
        print(f"[{time.strftime('%H:%M:%S')}] Cập nhật thành công! Tổng số {len(processed_items)} dòng.")
        return True
    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {e}")
        return False
    finally:
        if wb:
            try: wb.Close(False)
            except: pass
        if excel:
            try: excel.Quit()
            except: pass
        extract_lock.release()


class AutoUpdateHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=DIRECTORY, **kwargs)

    def end_headers(self):
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        super().end_headers()

    def do_GET(self):
        try:
            if self.path.startswith('/api/refresh'):
                success = extract_excel_data(force=True)
                self.send_response(200)
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                res = json.dumps({"status": "ok", "success": success}).encode('utf-8')
                self.wfile.write(res)
                return
            elif self.path in ('/', '/index.html') or self.path.startswith('/index.html'):
                extract_excel_data(force=True)
        except Exception as e:
            print(f"Lỗi trong do_GET: {e}")
        return super().do_GET()

if __name__ == '__main__':
    try:
        extract_excel_data()
    except Exception as e:
        print(f"Không thể đọc Excel lúc khởi động: {e}. Đang dùng dữ liệu đã cache...")
    
    socketserver.TCPServer.allow_reuse_address = True
    with socketserver.TCPServer(("", PORT), AutoUpdateHandler) as httpd:
        print(f"Server đang chạy tại http://localhost:{PORT}")
        print("Mỗi khi bạn mở hoặc F5 lại trang Dashboard, server sẽ tự kiểm tra và đọc file Excel mới nhất!")
        httpd.serve_forever()

